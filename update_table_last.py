import time
import warnings
from logging.handlers import RotatingFileHandler

import openpyxl
from dotenv import load_dotenv
from googleapiclient import discovery
from google.oauth2 import service_account
from googleapiclient.discovery import build
import logging
import os
import datetime as dt

from excel_editor import copy_excel, common_excel, del_start_excel

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs/')
log_file = os.path.join(BASE_DIR, 'logs/pars_stocks_table.log')
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler(
    log_file,
    maxBytes=100000,
    backupCount=3,
    encoding='utf-8'
)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s, %(levelname)s, %(message)s',
    handlers=(
        file_handler,
        console_handler
    )
)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
CREDENTIALS_FILE = 'credentials_service.json'
credentials = service_account.Credentials.from_service_account_file(CREDENTIALS_FILE)
service = discovery.build('sheets', 'v4', credentials=credentials)
dotenv_path = os.path.join(os.path.dirname(__file__), '.env')

if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
load_dotenv('.env ')
SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')
SPREADSHEET_ID_SALES = os.getenv('SPREADSHEET_ID_SALES')
NAME_SHEET = os.getenv('NAME_SHEET')


def convert_to_column_letter(column_number):
    column_letter = ''
    while column_number != 0:
        c = ((column_number - 1) % 26)
        column_letter = chr(c + 65) + column_letter
        column_number = (column_number - c) // 26
    return column_letter


def get_order_count(article_plan_fact):
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID_SALES,
                                range=f'{month}.{year}', majorDimension='ROWS').execute()
    values = result.get('values', [])
    if not values:
        logging.info('No data found.')
    else:
        for row in values[2:]:
            article = row[6]
            if int(article_plan_fact) == int(article):
                fbs = 14 + (int(day) - 1) * 6
                if row[fbs] == '':
                    row[fbs] = 0
                fbo = 15 + (int(day) - 1) * 6
                if row[fbo] == '':
                    row[fbo] = 0
                summ = int(row[fbs])+int(row[fbo])
                return summ


def update_table_order(table_id):
    date = f'{day}.{month}.{year}'
    article = ''
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    result = result.get('values', [])
    row = 3
    body_data = []
    if not result:
        logging.info('No data found.')
    else:
        for values in result:
            if date in values:
                column = values.index(date)+1
            for item in values:
                if item.startswith('Артикул'):
                    for i in item:
                        if i.isdigit():
                            article += i
                    value = get_order_count(article)
                    article = ''
                    try:
                        body_data += [
                            {'range': f'{range_name}!{convert_to_column_letter(column)}{row}',
                             'values': [[f'{value}']]}]
                        row += 13
                    except:
                        pass
                    finally:
                        body = {
                            'valueInputOption': 'USER_ENTERED',
                            'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def dicts_sales(employees_sheet):
    dict_sales = {}
    for x in range(2, employees_sheet.max_row + 1):
        article = str(employees_sheet.cell(row=x, column=9).value)
        if article not in dict_sales and article.isnumeric():
            dict_sales[article] = employees_sheet.cell(row=x, column=17).value
        elif article in dict_sales:
            dict_sales[article] += employees_sheet.cell(row=x, column=17).value
    return dict_sales


def update_table_sale(table_id, dicts_sales):
    date = f'{day}.{month}.{year}'
    article = ''
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    result = result.get('values', [])
    row = 4
    body_data = []
    if not result:
        logging.info('No data found.')
    else:
        for values in result:
            if date in values:
                column = values.index(date) + 1
            for item in values:
                if item.startswith('Артикул'):
                    for i in item:
                        if i.isdigit():
                            article += i
                    value = dicts_sales[article]
                    article = ''
                    try:
                        body_data += [
                            {'range': f'{range_name}!{convert_to_column_letter(column)}{row}',
                             'values': [[f'{value}']]}]
                        row += 13
                    except:
                        pass
                    finally:
                        body = {
                            'valueInputOption': 'USER_ENTERED',
                            'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()



if __name__ == '__main__':
    day = (dt.datetime.now() - dt.timedelta(days=1)).strftime('%d')
    month = (dt.datetime.now() - dt.timedelta(days=1)).strftime('%m')
    year = (dt.datetime.now() - dt.timedelta(days=1)).strftime("%Y")
    date_from = dt.datetime.date(dt.datetime.now())
    range_name = NAME_SHEET
    update_table_order(SPREADSHEET_ID)
    time.sleep(10)
    copy_excel(path='/home/werocket/wb/parser_fbo/excel_docs',path_new='/home/werocket/wb/plan_fact/excel_docs')
    time.sleep(1)
    path = '/home/werocket/wb/plan_fact/excel_docs'
    common_excel(path)
    time.sleep(1)
    del_start_excel(path)
    time.sleep(1)
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        excel_file = openpyxl.load_workbook(f'final_excel/ALL-{year}-{month}-{day}.xlsx')
    employees_sheet = excel_file['Sheet1']
    update_table_sale(SPREADSHEET_ID,dicts_sales(employees_sheet))



